import sys
import time
from PySide6 import QtCore, QtGui, QtWidgets
from PySide6.QtUiTools import QUiLoader
from PySide6.QtWidgets import QApplication, QMainWindow, QLineEdit, QFileDialog, QMessageBox, QTableWidgetItem, \
    QPushButton
from PySide6.QtCore import QDate, Qt
from PySide6.QtGui import QCloseEvent, QPen, QColor, QStandardItem, QStandardItemModel
from MainWindowui import Ui_MainWindow
# from docx import Document
# from docx.shared import Inches
import os
import subprocess
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Font, PatternFill, colors
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, FormulaRule
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
import xlsxwriter
import xlwings as xw
from xlwings.constants import InsertShiftDirection
from win32com.client import Dispatch

import pyvisa
import Calibrator
import Multimetr



# 1 instalacja PySide6 ... pip install PySide6 !
# 2 instalacja docx do Worda pip install python-docx... nie przyda sie chyba
# 3 aktualizacja gui pyside6-uic mainwindow.ui -o MainWindowui.py
# 4 instalacja Excela pip install openpyxl !
# 5 Import obrazu do excela pip install Pillow !
# 6 generowanie excela pip install xlsxwriter !
# pip install pandas !
# 7 pip install pywin32 and install packages Dispatch
# install package xlwings !
# 9 pip install aspose-cellsO JEDNAK NIE
# import pyvisa do komunikacji z urzadzeniami


if __name__ == "__main__":
    class MainWindow(QMainWindow, Ui_MainWindow):
        def __init__(self):
            super(MainWindow, self).__init__()
            self.setupUi(self)

            # cos do zapisu
            self.initUI()

            # aktualna data do SW
            self.data_wzorcowania.setDate(QDate.currentDate())

            # guziki do kalibratora
            self.wartosc_kalibrator.valueChanged.connect(self.odczytaj)
            self.Nastaw.clicked.connect(self.nastawa)
            self.Wyczysc.clicked.connect(self.wyczysc)

            # odczyt wartośći z listy AC/DC
            self.AC_DC.currentTextChanged.connect(self.text_changed)

            # usunąc potem
            self.AC_DC.currentTextChanged.connect(self.odczytaj)

            # odczyt wartości z listy mV/V/mA/A
            self.ustawienie_kalibrator.currentTextChanged.connect(self.text_changed)

            # usunąc potem
            self.ustawienie_kalibrator.currentTextChanged.connect(self.odczytaj)

            # podanie sciezki zapisu
            self.sciezkaSW_zapis.setText(u"Zapis_Swiadectw_Wzorcowania/nazwa_pliku")
            self.sciezkaWynik_zapis.setText(u"Zapis_Wynikow_Wzorcowania/Przyklad")
            self.sciezka_Model.setText(u"Modele/tescik")


            # odczyt wynikow
            self.odczyt_wynikow.clicked.connect(self.open_file_dialog)
            self.SzukajModelu.clicked.connect(self.open_file_dialog)

            # szukaj plikow
            self.SzukajWynikow.clicked.connect(self.save_file_dialog)
            self.SzukajSW.clicked.connect(self.save_file_dialog)


            # zamkniecie aplikacji
            self.zamkniecie_aplikacji.clicked.connect(QApplication.instance().quit)

            # wybor zglaszajacego
            # self.Wybor_zglaszajacy.currentTextChanged.connect(self.zglaszajacy)
            self.filled = False
            self.zglaszajacy()

            # generowanie swiadectwa
            self.generuj_swiadectwo.clicked.connect(self.swiadectwo_final)

            # zapisywanie wynikow
            # self.Zapisz_wynik.clicked.connect(self.save_to_excel)
            self.Zapisz_wynik.clicked.connect(self.update_excel_click)

            # Tworzenie excela
            self.Stworz.clicked.connect(self.generate_excel)

            # Zapisywanie pomiarow
            self.ZapiszDCV.clicked.connect(self.update_excel_click)
            self.ZapiszACV.clicked.connect(self.update_excel_click)
            self.ZapiszDCI.clicked.connect(self.update_excel_click)
            self.ZapiszACI.clicked.connect(self.update_excel_click)
            self.ZapiszR.clicked.connect(self.update_excel_click)

            # Wczytywanie modeli z listy
            self.wybierz_model.currentTextChanged.connect(self.load_model)
            # self.wybierz_model.currentIndexChanged.connect(self.load_model)

            # zmienianie zakladek
            self.tabWidget.currentChanged.connect(self.on_tab_changed)
            self.Tabwybor.currentChanged.connect(self.on_tab_changed2)

            # Podswietlenie tabeli

            # self.wynikiDCV.selectionModel().selectionChanged.connect(self.highlight_current_cell)
            self.count = 0
            self.NextDCV.clicked.connect(self.highlight_current_cell_click)
            self.NextACV.clicked.connect(self.highlight_current_cell_click)
            self.NextDCI.clicked.connect(self.highlight_current_cell_click)
            self.NextACI.clicked.connect(self.highlight_current_cell_click)
            # self.NextR.clicked.connect(self.highlight_current_cell_click)

            # Hovery
            self.zakresDCV.setToolTip('Pierwszy zakres')
            self.zakresDCV_2.setToolTip('Drugi zakres')
            self.zakresACV.setToolTip('Pierwszy zakres')
            self.zakresACV_2.setToolTip('Drugi zakres')
            self.zakresDCI.setToolTip('Pierwszy zakres')
            self.zakresDCI_2.setToolTip('Drugi zakres')
            self.zakresACI.setToolTip('Pierwszy zakres')
            self.zakresACI_2.setToolTip('Drugi zakres')

            self.ilDCV.setToolTip('Ilosc zakresow')
            self.ilACV.setToolTip('Ilosc zakresow')
            self.ilDCI.setToolTip('Ilosc zakresow')
            self.ilACI.setToolTip('Ilosc zakresow')
            self.ilR.setToolTip('Ilosc zakresow')


            # Pomiary

            self.initialize = False
            self.displayed_warning = False
            self.displayed_warningmA = False
            self.PomiarDCV.clicked.connect(self.pomiary)
            self.PomiarDCI.clicked.connect(self.pomiary)
            self.PomiarACV.clicked.connect(self.pomiary)
            self.PomiarACI.clicked.connect(self.pomiary)
            # self.PomiarR.clicked.connect(self.pomiary)



            # Kasuj wyniki
            self.KasujDCV.clicked.connect(self.clear_table)
            self.KasujACV.clicked.connect(self.clear_table)
            self.KasujDCI.clicked.connect(self.clear_table)
            self.KasujACI.clicked.connect(self.clear_table)
            self.KasujR.clicked.connect(self.clear_table)

            # Reset all
            self.Reset.clicked.connect(self.reset)

            # uzupelnianie comboboxow
            self.fil_modele()
            self.fill_adress()
            self.odswiez.clicked.connect(self.fill_adress)

            #komunikacja
            self.connect_kal()
            self.polacz_kal.clicked.connect(self.connect_kal)
            self.connect_dmm()
            self.polacz_dmm.clicked.connect(self.connect_dmm)


        # cos do zapisu
        def initUI(self):
            # ...
            self.saveFileDialog = QFileDialog()
            self.saveFileDialog.setAcceptMode(QFileDialog.AcceptSave)
            self.saveFileDialog.setFileMode(QFileDialog.AnyFile)
            self.saveFileDialog.setNameFilter("Excel Files (*.xlsx)")

        # zamkniecie aplikacji
        def closeEvent(self, event: QCloseEvent):
            czy_zamknac = QMessageBox.question(self, "Zamknąć aplikacje?", "Czy na pewno chcesz wyjśc z programu?",
                                               QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if czy_zamknac == QMessageBox.StandardButton.Yes:
                event.accept()
            else:
                event.ignore()

        def highlight_current_cell_click(self):

            self.count += 1

            sender = self.sender()
            if sender.objectName() == "NextDCV":
                self.highlight_current_cell(self.wynikiDCV)
            elif sender.objectName() == "NextACV":
                self.highlight_current_cell(self.wynikiACV)
            elif sender.objectName() == "NextDCI":
                self.highlight_current_cell(self.wynikiDCI)
            elif sender.objectName() == "NextACI":
                self.highlight_current_cell(self.wynikiACI)

        def highlight_current_cell(self, table):

            sender = self.sender()

            if self.initialize:

                if self.count == 1:
                    self.zakres = table.item(4, 1).text()

                for i in range(table.rowCount()):
                    for j in range(table.columnCount()):
                        item = table.item(i, j)
                        if item is not None:
                            item.setFlags(item.flags() | QtCore.Qt.ItemIsEditable)

                try:
                    move = table.currentRow()
                    table.selectRow(move + 1)
                    item = table.item(table.currentRow(), 2)
                    itemp = table.item(table.currentRow(), 1)
                    table.setCurrentCell(table.currentRow(), 2)

                    for i in range(0, 25):
                        if item.text() == '':
                            table.selectRow(move + 2)
                            item = table.item(table.currentRow(), 2)
                            table.setCurrentCell(table.currentRow(), 2)
                        # print(f"Teraz mamy wartosc: {item.text()}")
                        if itemp.text() == 'uV':
                            self.zakres = "uV"
                            print(f"Ustawiam wartosc na: {itemp.text()}")
                            self.calibrator_nastawa(item, self.zakres, sender.objectName())
                            self.update_multimetr(item, table)
                            break
                        elif itemp.text() == 'mV':
                            self.zakres = "mV"
                            print(f"Ustawiam wartosc na: {itemp.text()}")
                            self.calibrator_nastawa(item, self.zakres, sender.objectName())
                            self.update_multimetr(item, table)
                            break
                        elif itemp.text() == 'V':
                            self.zakres = "V"
                            print(f"Ustawiam wartosc na: {itemp.text()}")
                            self.calibrator_nastawa(item, self.zakres, sender.objectName())
                            self.update_multimetr(item, table)
                            break
                        elif itemp.text() == 'uA':
                            self.zakres = "uA"
                            print(f"Ustawiam wartosc na: {itemp.text()}")
                            self.calibrator_nastawa(item, self.zakres, sender.objectName())
                            self.update_multimetr(item, table)
                            break
                        elif itemp.text() == 'mA':
                            self.zakres = "mA"
                            if float(item.text()) > 400:
                                if self.displayed_warningmA == False:
                                    reply = QMessageBox.question(self, 'Zmień przewody',
                                                                 'Przekroczyłeś wartość 400mA. Czy zmieniłeś przewody?',
                                                                 QMessageBox.Yes | QMessageBox.No)
                                    if reply == QMessageBox.Yes:
                                        self.displayed_warningmA = True
                                        self.calibrator_nastawa(item, self.zakres, sender.objectName())
                                        self.update_multimetr(item, table)
                                    else:
                                        self.displayed_warningmA = False
                                        table.setCurrentCell(table.currentRow() - 1, 2)
                                else:
                                    print(f"Ustawiam wartosc na: {itemp.text()}")
                                    self.calibrator_nastawa(item, self.zakres, sender.objectName())
                                    self.update_multimetr(item, table)
                            break
                        elif itemp.text() == 'A':
                            self.zakres = "A"
                            if self.displayed_warningmA is False:
                                self.fluke5100b.write('S')
                                reply = QMessageBox.question(self, 'Zmień przewody',
                                                             'Przekroczyłeś wartość 400mA. Czy zmieniłeś przewody?',
                                                             QMessageBox.Yes | QMessageBox.No)
                                if reply == QMessageBox.Yes:
                                    self.displayed_warningmA = True
                                    self.calibrator_nastawa(item, self.zakres, sender.objectName())
                                    self.update_multimetr(item, table)
                                else:
                                    self.displayed_warningmA = False
                                    table.setCurrentCell(table.currentRow() - 1, 2)
                                    item = table.item(table.currentRow(), 2)
                                    if item.text() == '':
                                        table.setCurrentCell(table.currentRow() - 1, 2)
                                        itemp = table.item(table.currentRow() - i, 1)
                                        for j in range(0, 25):
                                            if itemp.text() != 'mA':
                                                itemp = table.item(table.currentRow() - j, 1)
                                            elif itemp.text() == 'mA':
                                                self.zakres = "mA"
                                                self.calibrator_nastawa(item, self.zakres, sender.objectName())
                                                self.update_multimetr(item, table)
                                                break
                            else:
                                print(f"Ustawiam wartosc na: {itemp.text()}")
                                self.calibrator_nastawa(item, self.zakres, sender.objectName())
                                self.update_multimetr(item, table)
                            try:
                                if float(item.text()) > 2:
                                    if self.displayed_warning is False:
                                        self.fluke5100b.write('S')
                                        reply = QMessageBox.question(self, 'Zmień przewody',
                                                                     'Przekroczyłeś wartość 2. Czy zmieniłeś przewody?',
                                                                     QMessageBox.Yes | QMessageBox.No)
                                        if reply == QMessageBox.Yes:
                                            self.displayed_warning = True
                                            self.calibrator_nastawa(item, self.zakres, sender.objectName())
                                        else:
                                            self.displayed_warning = False
                                            table.setCurrentCell(table.currentRow()-1, 2)
                                            item = table.item(table.currentRow(), 2)
                                            if item.text() == '':
                                                table.setCurrentCell(table.currentRow() - 1, 2)
                                    else:
                                        print(f"Ustawiam wartosc na: {itemp.text()}")
                                        self.calibrator_nastawa(item, self.zakres, sender.objectName())
                                        self.update_multimetr(item, table)
                            except ValueError:
                                pass
                            break
                        else:
                            itemp = table.item(table.currentRow() - i, 1)

                    # if self.displayed_warning is True and self.displayed_warningmA is True:
                    item = table.item(table.currentRow(), 2)
                    print(f"Teraz mamy wartosc: {item.text()} {self.zakres}")
                    self.odczyt_kalibrator.setText(f"{item.text()} {self.zakres}")

                    # self.fluke5100b.write(f'{item.text()}')

                    # self.calibrator_nastawa(item, self.zakres, None)

                    # ustawienie atrybutu ItemIsEditable na False dla wszystkich komórek
                    for i in range(table.rowCount()):
                        for j in range(table.columnCount()):
                            if i != table.currentRow() or j != 4:
                                # print(f"{i}{j}")
                                # print(table.currentRow())
                                item = table.item(i, j)
                                if item is not None:
                                    item.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable)
                            else:
                                table.setCurrentCell(i, j)
                                table.editItem(table.item(i, j))

                     # ustawienie atrybutu ItemIsEditable na True dla wybranej komórki
                    # it = QtWidgets.QTableWidgetItem("not editable")
                    # it.setFlags(it.flags() & ~QtCore.Qt.ItemIsEditable)
                    # table.setItem(1, 1, it)

                except AttributeError as e:
                    self.error1.setText(str(e))
                    print(e)
                    pass

        def calibrator_nastawa(self, item, zakres, sender):

            try:
                self.fluke5100b.write('CC')

                if zakres == "uV":
                    if sender == "PomiarDCV" or sender == "NextDCV":
                        self.fluke5100b.write(f'{item.text()}E-6V,')
                        self.fluke5100b.write('N')
                    elif sender == "PomiarACV" or sender == "NextACV":
                        self.fluke5100b.write(f'{item.text()}E-6V60H,')
                        self.fluke5100b.write('N')
                elif zakres == "mV":
                    if sender == "PomiarDCV" or sender == "NextDCV":
                        self.fluke5100b.write(f'{item.text()}E-3V,')
                        self.fluke5100b.write('N')
                    elif sender == "PomiarACV" or sender == "NextACV":
                        self.fluke5100b.write(f'{item.text()}E-3V60H,')
                        self.fluke5100b.write('N')
                elif zakres == "V":
                    if sender == "PomiarDCV" or sender == "NextDCV":
                        self.fluke5100b.write(f'{item.text()}V,')
                        self.fluke5100b.write('N')
                    elif sender == "PomiarACV" or sender == "NextACV":
                        self.fluke5100b.write(f'{item.text()}V60H,')
                        self.fluke5100b.write('N')
                elif zakres == "uA":
                    if sender == "PomiarDCI" or sender == "NextDCI":
                        self.fluke5100b.write(f'{item.text()}E-6A,')
                        self.fluke5100b.write('N')
                    elif sender == "PomiarACI" or sender == "NextACI":
                        self.fluke5100b.write(f'{item.text()}E-6A60H,')
                        self.fluke5100b.write('N')
                elif zakres == "mA":
                    if sender == "PomiarDCI" or sender == "NextDCI":
                        self.fluke5100b.write(f'{item.text()}E-3A,')
                        self.fluke5100b.write('N')
                    elif sender == "PomiarACI" or sender == "NextACI":
                        self.fluke5100b.write(f'{item.text()}E-3A60H,')
                        self.fluke5100b.write('N')
                elif zakres == "A":
                    if sender == "PomiarDCI" or sender == "NextDCI":
                        self.fluke5100b.write(f'{item.text()}A,')
                        self.fluke5100b.write('N')
                    elif sender == "PomiarACI" or sender == "NextACI":
                        self.fluke5100b.write(f'{item.text()}V60H,')
                        self.fluke5100b.write('N')
            except Exception as e:
                self.error3.setText(str(e))
                print(e)
                pass

        def calibrator_nastawa_value(self, item, zakres, sender):

            try:
                if zakres == "uV":
                    if self.AC_DC.currentText() == "DC" or sender == "PomiarDCV":
                        self.fluke5100b.write(f'{item.value()}E-6V,')
                        self.fluke5100b.write('N')
                    elif self.AC_DC.currentText() == "AC" or sender == "PomiarACV":
                        self.fluke5100b.write(f'{item.value()}E-6V60H,')
                        self.fluke5100b.write('N')
                elif zakres == "mV":
                    if self.AC_DC.currentText() == "DC" or sender == "PomiarDCV":
                        self.fluke5100b.write(f'{item.value()}E-3V,')
                        self.fluke5100b.write('N')
                    elif self.AC_DC.currentText() == "AC" or sender == "PomiarACV":
                        self.fluke5100b.write(f'{item.value()}E-3V60H,')
                        self.fluke5100b.write('N')
                elif zakres == "V":
                    if self.AC_DC.currentText() == "DC" or sender == "PomiarDCV":
                        self.fluke5100b.write(f'{item.value()}V,')
                        self.fluke5100b.write('N')
                    elif self.AC_DC.currentText() == "AC" or sender == "PomiarACV":
                        self.fluke5100b.write(f'{item.value()}V60H,')
                        self.fluke5100b.write('N')
                elif zakres == "uA":
                    if self.AC_DC.currentText() == "DC" or sender == "PomiarDCI":
                        self.fluke5100b.write(f'{item.value()}E-6A,')
                        self.fluke5100b.write('N')
                    elif self.AC_DC.currentText() == "AC" or sender == "PomiarACI":
                        self.fluke5100b.write(f'{item.value()}E-6A60H,')
                        self.fluke5100b.write('N')
                elif zakres == "mA":
                    if self.AC_DC.currentText() == "DC" or sender == "PomiarDCI":
                        self.fluke5100b.write(f'{item.value()}E-3A,')
                        self.fluke5100b.write('N')
                    elif self.AC_DC.currentText() == "AC" or sender == "PomiarACI":
                        self.fluke5100b.write(f'{item.value()}E-3A60H,')
                        self.fluke5100b.write('N')
                elif zakres == "A":
                    if self.AC_DC.currentText() == "DC" or sender == "PomiarDCI":
                        self.fluke5100b.write(f'{item.value()}A,')
                        self.fluke5100b.write('N')
                    elif self.AC_DC.currentText() == "AC" or sender == "PomiarACI":
                        self.fluke5100b.write(f'{item.value()}V60H,')
                        self.fluke5100b.write('N')
            except Exception as e:
                self.error3.setText(str(e))
                print(e)
                pass

        def pomiary(self):

            self.initialize = True
            self.displayed_warning = False
            self.displayed_warningmA = False
            self.count = 0

            try:
                # Reset -go to remote
                self.fluke5100b.write('J')
                self.fluke5100b.write('CC')
            except Exception as e:
                self.error1.setText(str(e))
                print(e)
                pass

            sender = self.sender()
            if sender.objectName() == "PomiarDCV":
                print("Teraz mierzymy DCV: ")
                self.wynikiDCV.setCurrentCell(5, 2)
                item = self.wynikiDCV.item(5, 2)
                itemp = self.wynikiDCV.item(4, 1)
                print(f"Mam wartosc: {item.text()} {itemp.text()}")
                self.calibrator_nastawa(item, itemp.text(), sender.objectName())
                time.sleep(self.timesleep.value())
                self.update_multimetr(item, self.wynikiDCV)
                self.blokuj(self.wynikiDCV)
            elif sender.objectName() == "PomiarACV":
                print("Teraz mierzymy ACV: ")
                self.wynikiACV.setCurrentCell(5, 2)
                item = self.wynikiACV.item(5, 2)
                itemp = self.wynikiACV.item(4, 1)
                print(f"Mam wartosc: {item.text()} {itemp.text()}")
                self.calibrator_nastawa(item, itemp.text(), sender.objectName())
                time.sleep(self.timesleep.value())
                self.update_multimetr(item, self.wynikiACV)
                self.blokuj(self.wynikiACV)
            elif sender.objectName() == "PomiarDCI":
                print("Teraz mierzymy DCI: ")
                self.wynikiDCI.setCurrentCell(5, 2)
                item = self.wynikiDCI.item(5, 2)
                itemp = self.wynikiDCI.item(4, 1)
                print(f"Mam wartosc: {item.text()} {itemp.text()}")
                self.calibrator_nastawa(item, itemp.text(), sender.objectName())
                time.sleep(self.timesleep.value())
                self.update_multimetr(item, self.wynikiDCI)
                self.blokuj(self.wynikiDCI)
            elif sender.objectName() == "PomiarACI":
                print("Teraz mierzymy ACI: ")
                self.wynikiACI.setCurrentCell(5, 2)
                item = self.wynikiACI.item(5, 2)
                itemp = self.wynikiACI.item(4, 1)
                print(f"Mam wartosc: {item.text()} {itemp.text()}")
                self.calibrator_nastawa(item, itemp.text(), sender.objectName())
                time.sleep(self.timesleep.value())
                self.update_multimetr(item, self.wynikiACI)
                self.blokuj(self.wynikiACI)
            elif sender.objectName() == "PomiarR":
                print("Teraz mierzymy R: ")
                self.wynikiR.setCurrentCell(5, 2)
                item = self.wynikiR.item(5, 2)
                itemp = self.wynikiR.item(4, 1)
                print(f"Mam wartosc: {item.text()} {itemp.text()}")
                self.calibrator_nastawa(item, itemp.text(), sender.objectName())

            # self.fluke5100b.write(f'{item.text()}')

        def blokuj(self, table):

            # ustawienie atrybutu ItemIsEditable na False dla wszystkich komórek
            for i in range(table.rowCount()):
                for j in range(table.columnCount()):
                    if i != table.currentRow() or j != 4:
                        # print(f"{i}{j}")
                        # print(table.currentRow())
                        item = table.item(i, j)
                        if item is not None:
                            item.setFlags(item.flags() & ~QtCore.Qt.ItemIsEditable)
                    else:
                        table.setCurrentCell(i, j)
                        table.editItem(table.item(i, j))

        def zglaszajacy(self):

            # print(self.Wybor_zglaszajacy.currentText())
            # if self.Wybor_zglaszajacy.currentText() == "Linetech":
            #     lt = "LINETECH S.A. \nul. Warecka 11A \n00-034 Warszawa"
            #     self.Zglaszajacy.setText(lt)
            # elif self.Wybor_zglaszajacy.currentText() == "Hitachi":
            #     hitachi = "Hitachi Energy Poland Sp. z o.o. \nul. Leszno 59 \n06-300 Przasnysz"
            #     self.Zglaszajacy.setText(hitachi)
            # elif self.Wybor_zglaszajacy.currentText() == "Hanza":
            #     hanza = "HANZA POLAND SP.Z O.O. \nAL.JEROZOLIMSKIE 38 \n56 - 120 BRZEG DOLNY"
            #     self.Zglaszajacy.setText(hanza)

            try:
                if self.filled is False:
                    df = pd.read_excel("Klienci/Baza_klientow.xlsx")
                    if df.size == 0:
                        return

                    # pobranie kolumny z nazwami klientów i zapisanie do listy
                    nazwy_klientow = df.iloc[:, 1].tolist()

                    # pobranie kolumny z danymi klientów i zapisanie do listy
                    dane_klientow = df.iloc[:, 2].tolist()

                    self.Wybor_zglaszajacy.addItems(nazwy_klientow)
                    self.Zglaszajacy.setText(dane_klientow[0])

                    self.filled = True

                # po wybraniu pozycji w comboboxie wyświetlanie danych klienta w textedit
                self.Wybor_zglaszajacy.currentIndexChanged.connect(lambda index:
                                                                   self.Zglaszajacy.setText(dane_klientow[index]))

            except Exception as e:
                self.error2.setText(str(e))
                print(e)
                pass

        def fil_modele(self):

            # self.wybierz_model.clear()

            # określenie ścieżki do folderu
            folder_path = 'Modele'

            # pobranie nazw plików z folderu
            file_names = os.listdir(folder_path)
            file_names = sorted(file_names)

            # dodanie nazw plików do QComboBox
            for file_name in file_names:
                filename_without_ext, ext = os.path.splitext(file_name)
                self.wybierz_model.addItem(filename_without_ext)

        def fill_adress(self):

            try:
                # obiekt VISA ResourceManager, który będzie zarządzał dostępem do urządzeń VISA.
                rm = pyvisa.ResourceManager()
                devices = rm.list_resources()
                print(devices)
                self.kalibrator_adres.clear()
                self.kalibrator_adres.addItems(devices)
                self.multimetr_adres.clear()
                self.multimetr_adres.addItems(devices)
                self.multimetr_adres.setCurrentIndex(1)
            except Exception as e:
                self.status_dmm.setText(str(e))
                self.status_kal.setText(str(e))
                print(e)
                pass

        def connect_kal(self):

            try:
                # obiekt VISA ResourceManager, który będzie zarządzał dostępem do urządzeń VISA.
                rm = pyvisa.ResourceManager()

                # adres GPIB kalibratora
                fluke5100b_address = self.kalibrator_adres.currentText()

                # Następnie otwórz połączenie z kalibratorem za pomocą metody "open_resource" obiektu ResourceManager,
                # podając jako argument adres GPIB kalibratora.
                self.fluke5100b = rm.open_resource(fluke5100b_address)

                ident = self.fluke5100b.query("?")
                self.status_kal.setText(f"Połączenie z urządzeniem {ident.strip()} zostało nawiązane.")

            except Exception as e:
                self.status_kal.setText(f"Wystąpił błąd podczas nawiązywania połączenia: {str(e)}")

        def connect_dmm(self):

            try:
                # obiekt VISA ResourceManager, który będzie zarządzał dostępem do urządzeń VISA.
                rm = pyvisa.ResourceManager()

                # adres GPIB kalibratora
                multimetr_address = self.multimetr_adres.currentText()

                # Następnie otwórz połączenie z kalibratorem za pomocą metody "open_resource" obiektu ResourceManager,
                # podając jako argument adres GPIB kalibratora.
                self.multimetr = rm.open_resource(multimetr_address)

                ident = self.multimetr.query("*IDN?")
                self.status_dmm.setText(f"Połączenie z urządzeniem {ident.strip()} zostało nawiązane.")

            except Exception as e:
                self.status_dmm.setText(f"Wystąpił błąd podczas nawiązywania połączenia: {str(e)}")

        def text_changed(self, s):  # s is a str

            if s == 'A':
                print(s)
                self.wartosc_kalibrator.setMaximum(2)
            else:
                self.wartosc_kalibrator.setMaximum(1000)

        def nastawa(self):

            try:
                self.fluke5100b.write('CC')
            except Exception as e:
                self.error1.setText(str(e))
                print(e)
                pass

            reply = QMessageBox.question(self, 'Zmień przewody',
                                         'Czy na pewno zmieniłeś przewody?',
                                         QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:

                print(f'Ustawiona wartosc kalibratora to: ', self.wartosc_kalibrator.value(),
                      self.ustawienie_kalibrator.currentText(), self.AC_DC.currentText())

                # self.odczyt_kalibrator.setText(str(self.wartosc_kalibrator.value()) + " " +
                #                                str(self.ustawienie_kalibrator.currentText()) + " " +
                #                                str(self.AC_DC.currentText()))

                self.calibrator_nastawa_value(self.wartosc_kalibrator, self.ustawienie_kalibrator.currentText(), None)

                va = self.ustawienie_kalibrator.currentText()

                try:
                    self.multimetr.timeout = 5000
                    time.sleep(self.timesleep.value())
                    if self.AC_DC.currentText() == "AC" and "V" in va:
                        response = self.multimetr.query('MEASure:VOLTage:AC?')
                    elif self.AC_DC.currentText() == "DC" and "V" in va:
                        response = self.multimetr.query('MEASure:VOLTage:DC?')
                    elif self.AC_DC.currentText() == "AC" and "A" in va:
                        response = self.multimetr.query('MEASure:CURRent:AC?')
                    elif self.AC_DC.currentText() == "DC" and "A" in va:
                        response = self.multimetr.query('MEASure:CURRent:DC?')
                    else:
                        print("Coś nie tak :/")
                        response = "Coś nie tak :/"
                    result = float(response)
                    if result < 1:
                        result = result * 1000
                        result = np.around(float(result), decimals=4)
                    elif result < 0.001:
                        result = result * 1000000
                        result = np.around(float(result), decimals=4)
                    else:
                        result = np.around(float(response), decimals=4)

                    self.odczyt_kalibrator.setText(str(result))
                except Exception as e:
                    self.error3.setText(str(e))
                    print(e)
                    pass
            else:
                print("Nie zmieniles przewodow")

        def odczytaj(self):
            print("Ustawiona wartosc kalibratora:", self.wartosc_kalibrator.value())
            # self.odczyt_kalibrator.setText(str(self.wartosc_kalibrator.value()) + " " +
            #                                str(self.ustawienie_kalibrator.currentText()) + " " +
            #                                str(self.AC_DC.currentText()))

        def wyczysc(self):

            try:
                self.fluke5100b.write('CC')
                self.wartosc_kalibrator.setValue(0)
            except Exception as e:
                self.error1.setText(str(e))
                print(e)
                pass


        # funkcja do otwierania folderu
        def open_folder(self, file_path):
            folder_path = os.path.dirname(file_path)  # uzyskanie ścieżki do folderu
            os.startfile(folder_path)  # otwarcie folderu w eksploratorze plików

        # funkcja do otwierania okna dialogowego
        def open_file_dialog(self):
            options = QFileDialog.Options()
            file_name, _ = QFileDialog.getOpenFileName(None, "Otwórz plik", "",
                                                       "All Files (*);;Excel Files (*.xlsx)", options=options)
            if file_name:
                sender = self.sender()
                # tutaj ustawiam ścieżkę na QLineEdit
                if sender.objectName() == "SzukajSW":
                    path = file_name
                    filename_without_ext, ext = os.path.splitext(path)
                    self.sciezkaSW_zapis.setText(filename_without_ext)
                    print(file_name)
                elif sender.objectName() == "SzukajModelu":
                    path = file_name
                    filename_without_ext, ext = os.path.splitext(path)
                    if self.check_dcv.isChecked():
                        self.loadExcelData2(path, self.wynikiDCV, "DCV")
                    if self.check_acv.isChecked():
                        self.loadExcelData2(path, self.wynikiACV, "ACV")
                    if self.check_dci.isChecked():
                        self.loadExcelData2(path, self.wynikiDCI, "DCI")
                    if self.check_aci.isChecked():
                        self.loadExcelData2(path, self.wynikiACI, "ACI")
                    if self.check_r.isChecked():
                        self.loadExcelData2(path, self.wynikiR, "R")
                        self.sciezka_Model.setText(filename_without_ext)
                    print(file_name)
                elif sender.objectName() == "SzukajWynikow" or "odczyt_wynikow":
                    path = file_name
                    filename_without_ext, ext = os.path.splitext(path)
                    self.sciezkaWynik_zapis.setText(filename_without_ext)
                    self.loadExcelData(self.sciezkaWynik_zapis.text() + ".xlsx", self.wyniki_wzorcowania)
                    print(file_name)

        def worksheet_exists(self, workbook_path, sheet_name):
            workbook = load_workbook(workbook_path)
            sheet_names = workbook.sheetnames
            return sheet_name in sheet_names

        def loadExcelData(self, path, table):

            table.clear()
            # workbook = load_workbook(path)

            if self.worksheet_exists(path, "R"):
                df5 = pd.read_excel(path, sheet_name="R")
            else:
                df5 = pd.DataFrame()
            if self.worksheet_exists(path, "ACI"):
                df4 = pd.read_excel(path, sheet_name="ACI")
            else:
                df4 = pd.DataFrame()
            if self.worksheet_exists(path, "DCI"):
                df3 = pd.read_excel(path, sheet_name="DCI")
            else:
                df3 = pd.DataFrame()
            if self.worksheet_exists(path, "ACV"):
                df2 = pd.read_excel(path, sheet_name="ACV")
            else:
                df2 = pd.DataFrame()
            if self.worksheet_exists(path, "DCV"):
                df1 = pd.read_excel(path, sheet_name="DCV")
            else:
                df1 = pd.DataFrame()

            df = pd.concat([df1, df2, df3, df4, df5], ignore_index=True)

            x = 10  # zmienic to na ta sama wartosc co w generate excel 399, 740

            df.fillna('', inplace=True)
            table.setRowCount(df.shape[0]+5)
            table.setColumnCount(df.shape[1]+2)
            column_headers = df.iloc[3+x]
            table.setHorizontalHeaderLabels(column_headers)

            # ustawienie dopasowywania się rozmiaru kolumn
            for i in range(table.columnCount()):
                 table.resizeColumnToContents(i)

            # Wyświetlanie danych z Excela
            for row in df.iterrows():
                values = row[1]
                for col_index, value in enumerate(values):
                    tableItem = QTableWidgetItem(str(value))
                    table.setItem(row[0], col_index, tableItem)

            # wyswietlanie excela od 10 lini
            for i in range(x):
                table.removeRow(0)

            # Ustawienie poprakwi kolumn 2 i 3 w kolumnie 4
            table.itemChanged.connect(lambda item: self.update_tablewidget(item, table2=table))

            # self.update_tablewidget()
            table.setStyleSheet("QTableView::item { border: 0px solid black; }")

        def loadExcelData2(self, path, table, sheet):

            table.clear()

            df = pd.read_excel(path, sheet_name=sheet)
            if df.size == 0:
                return

            x = 10 # zmienic to na ta sama wartosc co w generate excel 366, 740

            df.fillna('', inplace=True)
            # table.setRowCount(df.shape[0]+5)
            table.setRowCount(70)
            table.setColumnCount(df.shape[1]+2)
            column_headers = df.iloc[3+x]
            table.setHorizontalHeaderLabels(column_headers)

            # ustawienie dopasowywania się rozmiaru kolumn
            for i in range(table.columnCount()):
                table.resizeColumnToContents(i)


            # Wyświetlanie danych z Excela
            for row in df.iterrows():
                values = row[1]
                for col_index, value in enumerate(values):
                    tableItem = QTableWidgetItem(str(value))
                    table.setItem(row[0], col_index, tableItem)

            # wyswietlanie excela od 10 lini
            for i in range(x):
                table.removeRow(0)

            # Ustawienie poprakwi kolumn 2 i 3 w kolumnie 4
            table.itemChanged.connect(lambda item: self.update_tablewidget(item, table2=table))

            # self.update_tablewidget()
            table.setStyleSheet("QTableView::item { border: 0px solid black; }")

        def on_tab_changed(self, index):
            # Wywołanie metody po zmianie aktywnej zakładki
            if index == 1:  # sprawdzanie czy zmieniona została zakładka 2
                self.loadExcelData(self.sciezka_Model.text() + ".xlsx", self.wyniki_wzorcowania)
                path = self.sciezka_Model.text() + ".xlsx"
                filename_without_ext, ext = os.path.splitext(path)
                self.sciezkaWynik_zapis.setText(filename_without_ext)

            try:
                self.fluke5100b.write('S')
            except Exception as e:
                self.error1.setText(str(e))
                print(e)
                pass

        def on_tab_changed2(self, index):
            # Wywołanie metody po zmianie aktywnej zakładki
            if index == 0:  # sprawdzanie czy zmieniona została zakładka 2
                print("Jestes w zakladce DCV")
                QMessageBox.information(self, "Uwaga", "Sprawdź podłączenie przewodów")
            if index == 1:  # sprawdzanie czy zmieniona została zakładka 2
                print("Jestes w zakladce ACV")
                QMessageBox.information(self, "Uwaga", "Sprawdź podłączenie przewodów")
            if index == 2:  # sprawdzanie czy zmieniona została zakładka 2
                print("Jestes w zakladce DCI")
                QMessageBox.information(self, "Uwaga", "Sprawdź podłączenie przewodów")
            if index == 3:  # sprawdzanie czy zmieniona została zakładka 2
                print("Jestes w zakladce ACI")
                QMessageBox.information(self, "Uwaga", "Sprawdź podłączenie przewodów")

            try:
                self.fluke5100b.write('S')
            except Exception as e:
                self.error1.setText(str(e))
                print(e)
                pass

        def load_model(self):

            try:
                path = (u"Modele/" + self.wybierz_model.currentText() + ".xlsx")
                if self.worksheet_exists(path, "DCV"):
                    self.loadExcelData2(path, self.wynikiDCV, "DCV")
                if self.worksheet_exists(path, "ACV"):
                    self.loadExcelData2(path, self.wynikiACV, "ACV")
                if self.worksheet_exists(path, "DCI"):
                    self.loadExcelData2(path, self.wynikiDCI, "DCI")
                if self.worksheet_exists(path, "ACI"):
                    self.loadExcelData2(path, self.wynikiACI, "ACI")
                if self.worksheet_exists(path, "R"):
                    self.loadExcelData2(path, self.wynikiR, "R")
                filename_without_ext, ext = os.path.splitext(path)
                self.sciezka_Model.setText(filename_without_ext)
                # print(f"Sciezka: {filename_without_ext}")
            except FileNotFoundError as e:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Critical)
                msg.setText("Nie ma takiego pliku")
                msg.setInformativeText(str(e))
                msg.setWindowTitle("Błąd")
                msg.exec()
            except ValueError:
                pass
            except Exception as e:
                self.error2.setText(str(e))
                print(e)
                pass

        def clear_table(self):

            sender = self.sender()
            if sender.objectName() == "KasujDCV":
                self.odblokuj(self.wynikiDCV)
                #self.wynikiDCV.clear()
            elif sender.objectName() == "KasujACV":
                self.odblokuj(self.wynikiACV)
                #self.wynikiACV.clear()
            elif sender.objectName() == "KasujDCI":
                self.odblokuj(self.wynikiDCI)
                #s elf.wynikiDCI.clear()
            elif sender.objectName() == "KasujACI":
                self.odblokuj(self.wynikiACI)
                # self.wynikiACI.clear()
            elif sender.objectName() == "KasujR":
                self.odblokuj(self.wynikiR)
                # self.wynikiR.clear()

        def odblokuj(self, table):

            for i in range(table.rowCount()):
                for j in range(table.columnCount()):
                    item = table.item(i, j)
                    if item is not None:
                        item.setFlags(item.flags() | QtCore.Qt.ItemIsEditable)

        def reset(self):


            try:
                # Reset -go to local
                self.fluke5100b.write('*')
            except Exception as e:
                self.error2.setText(str(e))
                print(e)
                pass

            self.wynikiDCV.clear()
            self.wynikiACV.clear()
            self.wynikiDCI.clear()
            self.wynikiACI.clear()
            self.wynikiR.clear()

            self.wartosc_kalibrator.setValue(0)

            self.count = 0
            self.initialize = False

        def update_tablewidget(self, item, table2):
            # Obliczanie poprakwi
            row = item.row()
            col = item.column()
            try:
                if col == 3 or col == 4:
                    value1 = float(table2.item(row, 3).text())
                    value2 = float(table2.item(row, 4).text())
                    result = np.around(value1 - value2, decimals=8)
                    table2.item(row, 5).setText(str(result))
            except ValueError:
                pass
            except AttributeError:
                pass
            try:
                if col == 1:
                    value_1 = float(float(table2.item(row, col).text()) * 0.1)
                    value_2 = float(float(table2.item(row, col).text()) * 0.5)
                    value_3 = float(float(table2.item(row, col).text()) * 0.9)
                    value_4 = float(float(table2.item(row, col).text()) * (-0.9))

                    table2.item(row, 2).setText(str(value_1))
                    table2.item(row+1, 2).setText(str(value_2))
                    table2.item(row+2, 2).setText(str(value_3))
                    table2.item(row+3, 2).setText(str(value_4))
            except ValueError:
                pass
            except AttributeError:
                pass

        def update_multimetr(self, item, table):

            # Obliczanie poprakwi
            row = item.row()
            col = item.column()
            try:
                time.sleep(self.timesleep.value())
                if table == self.wynikiDCV:
                    response = self.multimetr.query('MEASure:VOLTage:DC?')
                elif table == self.wynikiACV:
                    response = self.multimetr.query('MEASure:VOLTage:AC?')
                elif table == self.wynikiDCI:
                    response = self.multimetr.query('MEASure:CURRent:DC?')
                elif table == self.wynikiACI:
                    response = self.multimetr.query('MEASure:CURRent:AC?')
                else:
                    print("Coś nie tak :/")
                    response = 0
                print(f"Odpowiedz: {response}")
                result = float(response)
                print(f"To po konwercji: {result}")
                if result < 1:
                    result = result * 1000
                    result = np.around(float(result), decimals=4)
                    table.item(row, col + 1).setText(str(result))
                elif result < 0.001:
                    result = result * 1000000
                    result = np.around(float(result), decimals=4)
                    table.item(row, col + 1).setText(str(result))
                else:
                    result = np.around(float(response), decimals=4)
                    table.item(row, col + 1).setText(str(result))
            except ValueError as e:
                print(e)
                pass
            except AttributeError as e:
                print(e)
                pass


        def save_to_excel(self):

            # Tworzenie nowego arkusza w pliku Excel
            wb = openpyxl.Workbook()
            ws = wb.active

            # ustawianie stylu obramowania dla komórek
            border_style = Side(border_style='thin', color='000000')
            border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

            alignmentCC = Alignment(wrap_text=True, horizontal='center', vertical='center')

            # Pobieranie danych z QTableWidget i zapisywanie ich do arkusza
            for row in range(self.wyniki_wzorcowania.rowCount()):
                for col in range(self.wyniki_wzorcowania.columnCount()):
                    item = self.wyniki_wzorcowania.item(row, col)
                    if item is not None:
                        ws.cell(row=row + 2, column=col + 1, value=item.text())


            # ustawienie obramowania dla komórki i dopasowanie szerokosci kolumn
            for row in range(self.wyniki_wzorcowania.rowCount()):
                for col in range(self.wyniki_wzorcowania.columnCount()):
                    column_letter = get_column_letter(col + 1)
                    column_dimensions = ws.column_dimensions[column_letter]
                    max_length = 0
                    for cell in ws[column_letter]:
                        try:
                            cell_value = str(cell.value)
                        except:
                            cell_value = ""
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                            # print(max_length)
                    if max_length < 30:
                        adjusted_width = (max_length + 2)
                        column_dimensions.width = adjusted_width
                    item = self.wyniki_wzorcowania.item(row, col)
                    if item is not None:
                        if item.text() != "":
                            cell1 = ws.cell(row=row + 2, column=col + 1)
                            cell1.border = border
                            cell1.alignment = Alignment(wrapText=True)
                            cell1.alignment = alignmentCC

            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 13
            ws.column_dimensions['F'].width = 14

            self.zapis_pliku(self.sciezkaWynik_zapis.text(), wb)

        def update_excel_click(self):

            sender = self.sender()
            try:
                path = self.sciezka_Model.text() + ".xlsx"
                if sender.objectName() == "ZapiszDCV":
                    self.update_excel(path, "DCV", self.wynikiDCV)
                elif sender.objectName() == "ZapiszACV":
                    self.update_excel(path, "ACV", self.wynikiACV)
                elif sender.objectName() == "ZapiszDCI":
                    self.update_excel(path, "DCI", self.wynikiDCI)
                elif sender.objectName() == "ZapiszACI":
                    self.update_excel(path, "ACI", self.wynikiACI)
                elif sender.objectName() == "ZapiszR":
                    self.update_excel(path, "R", self.wynikiR)
                elif sender.objectName() == "Zapisz_wynik":
                    path = self.sciezkaWynik_zapis.text() + ".xlsx"
                    self.update_excel2(path, self.wyniki_wzorcowania)
            except FileNotFoundError:
                path = u"Modele/" + self.wybierz_model.currentText() + ".xlsx"
                if sender.objectName() == "ZapiszDCV":
                    self.update_excel3(path, "DCV", self.wynikiDCV)
                elif sender.objectName() == "ZapiszACV":
                    self.update_excel3(path, "ACV", self.wynikiACV)
                elif sender.objectName() == "ZapiszDCI":
                    self.update_excel3(path, "DCI", self.wynikiDCI)
                elif sender.objectName() == "ZapiszACI":
                    self.update_excel3(path, "ACI", self.wynikiACI)
                elif sender.objectName() == "ZapiszR":
                    self.update_excel3(path, "R", self.wynikiR)
                elif sender.objectName() == "Zapisz_wynik":
                    self.update_excel2(path, self.wyniki_wzorcowania)


        def update_excel(self, path, sheet, table):

            # Wczytywanie pliku excel
            wb = openpyxl.load_workbook(path)
            ws = wb[sheet]

            x = 10 #zmienic tez w innych przy mzianie tego


            for row in range(table.rowCount()):
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    try:
                        if item is not None:
                            ws.cell(row=row+2+x, column=col+1, value=item.text())
                    except AttributeError:
                        pass

            try:
                wb.save(path)
            except PermissionError as e:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Critical)
                msg.setText("Plik otwarty")
                msg.setInformativeText(str(e))
                msg.setWindowTitle("Błąd")
                msg.exec()

        def update_excel3(self, path, sheet, table):

            # Wczytywanie pliku excel
            wb = openpyxl.load_workbook(path)
            ws = wb[sheet]

            x = 10  # zmienic tez w innych przy mzianie tego

            for row in range(table.rowCount()):
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    try:
                        if item is not None:
                            ws.cell(row=row+2+x, column=col+1, value=item.text())
                    except AttributeError:
                        pass

            try:
                path2 = self.sciezka_Model.text() + ".xlsx"
                wb.save(path2)
            except PermissionError as e:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Critical)
                msg.setText("Plik otwarty")
                msg.setInformativeText(str(e))
                msg.setWindowTitle("Błąd")
                msg.exec()

        def update_excel2(self, path, table):

            # Wczytywanie pliku excel
            wb = openpyxl.load_workbook(path)
            ws = wb.active

            x = 10  # zmienic tez w innych przy mzianie tego

            ile = 0
            temp = 0
            access = 0

            for row in range(table.rowCount()):
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    if ile == 2 and access < 1:
                        ws.cell(row=row + 2 + x - temp, column=col, value="") #zeby nie bylo slowa zakres
                        ws = wb[str(wb.sheetnames[1])]
                        temp = row - 3
                        access += 1
                    elif ile == 3 and access < 2:
                        ws.cell(row=row + 2 + x - temp, column=col, value="")
                        ws = wb[str(wb.sheetnames[2])]
                        temp = row - 3
                        access += 1
                    elif ile == 4 and access < 3:
                        ws.cell(row=row + 2 + x- temp, column=col, value="")
                        ws = wb[str(wb.sheetnames[3])]
                        temp = row - 3
                        access += 1
                    elif ile == 5 and access < 4:
                        ws.cell(row=row + 2 + x - temp, column=col, value="")
                        ws = wb[str(wb.sheetnames[4])]
                        temp = row - 3
                        access += 1
                    try:
                        if item is not None:
                            ws.cell(row=row+2 + x - temp, column=col+1, value=item.text())
                            if col == 1 and item.text() == "Zakres":
                                ile += 1

                    except AttributeError:
                        pass

            try:
                path = self.sciezkaWynik_zapis.text() + ".xlsx"
                wb.save(path)
            except PermissionError as e:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Critical)
                msg.setText("Plik otwarty")
                msg.setInformativeText(str(e))
                msg.setWindowTitle("Błąd")
                msg.exec()


        def generate_excel(self):

            wb = Workbook()

            if self.check_r.isChecked():
                ws5 = wb.create_sheet("R", 0)
                self.creat_excel(ws5, self.check_r, self.ilR.value(),
                                 self.zakresR.value(), self.zakresR.value())
            if self.check_aci.isChecked():
                ws4 = wb.create_sheet("ACI", 0)
                self.creat_excel(ws4, self.check_aci, self.ilACI.value(),
                                 self.zakresACI.value(), self.zakresACI_2.value())
            if self.check_dci.isChecked():
                ws3 = wb.create_sheet("DCI", 0)
                self.creat_excel(ws3, self.check_dci, self.ilDCI.value(),
                                 self.zakresDCI.value(), self.zakresDCI_2.value())
            if self.check_acv.isChecked():
                ws2 = wb.create_sheet("ACV", 0)
                self.creat_excel(ws2, self.check_acv, self.ilACV.value(),
                                 self.zakresACV.value(), self.zakresACV_2.value())
            if self.check_dcv.isChecked():
                ws1 = wb.create_sheet("DCV", 0)
                self.creat_excel(ws1, self.check_dcv, self.ilDCV.value(),
                                 self.zakresDCV.value(), self.zakresDCV_2.value())

            self.zapis_pliku(self.sciezka_Model.text(), wb)

            wb.close()

            if self.check_dcv.isChecked():
                self.loadExcelData2(self.sciezka_Model.text() + ".xlsx", self.wynikiDCV, "DCV")
            if self.check_acv.isChecked():
                self.loadExcelData2(self.sciezka_Model.text() + ".xlsx", self.wynikiACV, "ACV")
            if self.check_dci.isChecked():
                self.loadExcelData2(self.sciezka_Model.text() + ".xlsx", self.wynikiDCI, "DCI")
            if self.check_aci.isChecked():
                self.loadExcelData2(self.sciezka_Model.text() + ".xlsx", self.wynikiACI, "ACI")
            if self.check_r.isChecked():
                self.loadExcelData2(self.sciezka_Model.text() + ".xlsx", self.wynikiR, "R")

        def creat_excel(self, ws, checked, ile, zakres1, zakres2):

            col_headers = ['Zakres', 'Punkt Pomiarowy', 'Wartość napięcia odniesienia',
                           'Zmierzona wartość napięcia', 'Poprawka', 'Niepewność pomiaru']

            # ustawianie stylu obramowania dla komórek
            border_style = Side(border_style='thin', color='000000')
            border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

            # ustawianie czcionki i pozycji tekstu
            alignmentCC = Alignment(wrap_text=True, horizontal='center', vertical='center')
            fontA10 = Font(name='Arial', size=10)

            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 13
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 11

            if checked.isChecked():
                x = 10 # 0, tutaj zmienie chyba na wiecej zmienic TEZ W 398, 366

                for i, col in enumerate(col_headers):
                    ws.cell(5 + x, i + 2, col)
                    cell = ws.cell(row=5 + x, column=i+2)
                    cell.border = border
                    cell.font = fontA10
                    cell.alignment = alignmentCC

                if (ile <= 4) and (zakres2 / 100 >= 1) or \
                   (ile == 3) and (zakres2 / 10 >= 1) or \
                   (ile <= 2) and (zakres2 > 1) and (zakres2 < 10):
                    row_range = ile * 4 + 1
                elif (5 <= ile <= 7) and (zakres2 < 1) or \
                     (ile == 6) and (zakres2 / 10 < 1) and (zakres2 > 1):
                    row_range = ile * 4 + 3
                else:
                    row_range = ile * 4 + 2
                if checked != self.check_r:
                    for row in range(row_range):
                        row = row + x
                        try:
                            if (row == 0 + x) and (ile <= 4) and (zakres2 / 100 >= 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 0 + x) and (5 <= ile <= 7) and (zakres2 / 100 >= 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'mV')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'mA')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 5 + x) and (ile == 5) and (zakres2 / 100 >= 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{11 + x}:G{11 + x}")
                            elif (row == 9 + x) and (ile == 6) and (zakres2 / 100 >= 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{15 + x}:G{15 + x}")
                            elif (row == 13 + x) and (ile == 7) and (zakres2 / 100 >= 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{19 + x}:G{19 + x}")
                            elif (row == 0 + x) and (ile <= 3) and (zakres2 / 10 >= 1) and (zakres2 / 100 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 0 + x) and (4 <= ile <= 6) and (zakres2 / 10 >= 1) and (zakres2 / 100 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'mV')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'mA')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 5 + x) and (ile == 4) and (zakres2 / 10 >= 1) and (zakres2 / 100 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{11 + x}:G{11 + x}")
                            elif (row == 9 + x) and (ile == 5) and (zakres2 / 10 >= 1) and (zakres2 / 100 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{15 + x}:G{15 + x}")
                            elif (row == 13 + x) and (ile == 6) and (zakres2 / 10 >= 1) and (zakres2 / 100 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{19 + x}:G{19 + x}")
                            elif (row == 0 + x) and (ile <= 2) and (zakres2 / 10 < 1) and (zakres2 > 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 0 + x) and (3 <= ile <= 5) and (zakres2 / 10 < 1) and (zakres2 > 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'mV')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'mA')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 5 + x) and (ile == 3) and (zakres2 / 10 < 1) and (zakres2 > 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{11 + x}:G{11 + x}")
                            elif (row == 9 + x) and (ile == 4) and (zakres2 / 10 < 1) and (zakres2 > 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{15 + x}:G{15 + x}")
                            elif (row == 13 + x) and (ile == 5) and (zakres2 / 10 < 1) and (zakres2 > 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{19 + x}:G{19 + x}")
                            elif (row == 0 + x) and (ile == 6) and (zakres2 / 10 < 1) and (zakres2 > 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'uV')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'uA')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 5 + x) and (ile == 6) and (zakres2 / 10 < 1) and (zakres2 > 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'mV')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'mA')
                                ws.merge_cells(f"B{11 + x}:G{11 + x}")
                            elif (row == 18 + x) and (ile == 6) and (zakres2 / 10 < 1) and (zakres2 > 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{24 + x}:G{24 + x}")
                            elif (row == 0 + x) and (ile <= 2) and (zakres2 / 10 < 1) and (zakres2 > 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 0 + x) and (ile <= 1) and (zakres2 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 0 + x) and (2 <= ile <= 4) and (zakres2 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'mV')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'mA')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 5 + x) and (ile == 2) and (zakres2 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{11 + x}:G{11 + x}")
                            elif (row == 9 + x) and (ile == 3) and (zakres2 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{15 + x}:G{15 + x}")
                            elif (row == 13 + x) and (ile == 4) and (zakres2 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{19 + x}:G{19 + x}")
                            elif (row == 0 + x) and (5 <= ile <= 6) and (zakres2 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'uV')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'uA')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 5 + x) and (ile == 5) and (zakres2 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'mV')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'mA')
                                ws.merge_cells(f"B{11 + x}:G{11 + x}")
                            elif (row == 9 + x) and (ile == 6) and (zakres2 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'mV')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'mA')
                                ws.merge_cells(f"B{15 + x}:G{15 + x}")
                            elif (row == 18 + x) and (ile == 5) and (zakres2 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{24 + x}:G{24 + x}")
                            elif (row == 22 + x) and (ile == 6) and (zakres2 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{28 + x}:G{28 + x}")
                            elif row < row_range - 22 + x:
                                if zakres2 / 10 < 1 < zakres2:
                                    ws.cell(row + 6, 2, zakres2 * 100)
                                elif zakres2 / 10 < 1:
                                    ws.cell(row + 6, 2, zakres2 * 100)
                                else:
                                    ws.cell(row + 6, 2, zakres2 / 10)
                            elif row < row_range - 17 + x:
                                if zakres2 < 1:
                                    ws.cell(row + 6, 2, zakres2 * 1000)
                                else:
                                    ws.cell(row + 6, 2, zakres2)
                            elif row < row_range - 13 + x:
                                if zakres2 / 100 < 1:
                                    ws.cell(row + 6, 2, zakres2 * 10)
                                elif zakres2 < 1:
                                    ws.cell(row + 6, 2, zakres2 * 100)
                                else:
                                    ws.cell(row + 6, 2, zakres2 / 100)
                            elif row < row_range - 10 + x:
                                if zakres2 / 10 < 1:
                                    ws.cell(row + 6, 2, zakres2 * 100)
                                else:
                                    ws.cell(row + 6, 2, zakres2 / 10)
                            elif row < row_range - 4 + x:
                                if zakres2 < 1:
                                    ws.cell(row + 6, 2, zakres2 * 1000)
                                else:
                                    ws.cell(row + 6, 2, zakres2)
                            elif row < row_range + 1 + x:
                                ws.cell(row + 6, 2, zakres1)
                        except AttributeError:
                            pass
                        for col in range(2, 8):
                            cell = ws.cell(row=row+6, column=col)
                            cell.border = border
                            cell.alignment = alignmentCC
                    if ile == 6 and zakres2 / 100 >= 1:
                        end = 1
                    elif ile == 5 and zakres2 / 100 >= 1:
                        end = 2
                    elif ile <= 4 and zakres2 / 100 >= 1:
                        end = 0
                    elif ile == 6 and zakres2 / 10 >= 1:
                        end = 0
                    elif ile == 5 and zakres2 / 10 >= 1:
                        end = 1
                    elif ile == 4 and zakres2 / 10 >= 1:
                        end = 2
                    elif ile == 3 and zakres2 / 10 >= 1:
                        end = 0
                    elif ile <= 4 and zakres2 / 10 >= 1:
                        end = 3
                    elif ile == 3 and zakres2 < 1:
                        end = 1
                    elif ile == 2 and zakres2 < 1:
                        end = 2
                    elif ile == 4 and zakres2 / 10 < 1 < zakres2:
                        end = 1
                    elif ile == 6 and zakres2 < 1:
                        end = 1
                    elif ile == 5 and zakres2 < 1:
                        end = 2
                    elif ile == 6 and zakres2 / 10 < 1 < zakres2:
                        end = 2
                    else:
                        end = 0

                    endy = 0
                    y = 4 * (3 - end)

                    if (ile <= 4) and (zakres2 / 100 >= 1) or \
                            (ile == 3) and (zakres2 / 10 >= 1) or \
                            (ile <= 2) and (zakres2 > 1) and (zakres2 < 10):
                        row_range = ile * 4 + 1
                        for row in range(row_range + 3):
                            row = row + x
                            if (row - (x + 7) >= 0) and (row - (x + 7)) % 4 == 0 and end < 3 or \
                                    (row - (x + 7) >= 0) and (row - (x + 7)) % 4 == 0 and ile == 4 and end <= 3:
                                end += 1
                                ws.merge_cells(f"B{row}:B{row + 3}")
                    elif (5 <= ile <= 7) and (zakres2 < 1) or \
                            (ile == 6) and (zakres2 / 10 < 1) and (zakres2 > 1):
                        row_range = ile * 4 + 3
                        for row in range(row_range + 3):
                            row = row + x
                            if (row - (x + 7) >= 0) and (row - (x + 7)) % 4 == 0 and end < 3 or \
                                    (row - (x + 11) >= 0) and (row - (x + 11)) % 4 == 0 and end < 3:
                                end += 1
                                ws.merge_cells(f"B{row}:B{row + 3}")
                            elif ((100 > row > 60 and row > (6 + x + y)) or 45 > row > (6 + x + y)) and \
                                    (row - x) % 4 == 0 and end >= 3 > endy or \
                                    row > (6 + x + y) and (row > 101 and (row - 102) % 4 == 0) and end >= 3 > endy or \
                                    (60 > row > 45 and (row - (42 + x)) % 4 == 0) and end >= 3 > endy:
                                endy += 1
                                ws.merge_cells(f"B{row}:B{row + 3}")
                            elif (row - (25 + x)) % 4 == 0 and row > y + 19 + x and endy >= 3:
                                ws.merge_cells(f"B{row}:B{row + 3}")
                    else:
                        row_range = ile * 4 + 2
                        for row in range(row_range + 3):
                            row = row + x
                            if (row - (x + 7) >= 0) and (row - (x + 7)) % 4 == 0 and end < 3 or \
                                    (row - (x + 11) >= 0) and (row - (x + 11)) % 4 == 0 and end < 3:
                                end += 1
                                ws.merge_cells(f"B{row}:B{row + 3}")
                            elif ((100 > row > 60 and row > (6 + x + y)) or 45 > row > (6 + x + y)) and \
                                    (row - x) % 4 == 0 and end >= 3 or \
                                    row > (6 + x + y) and (row > 101 and (row-102) % 4 == 0) and end >= 3 or \
                                    (60 > row > 45 and (row - (42 + x)) % 4 == 0) and end >= 3:
                                ws.merge_cells(f"B{row}:B{row + 3}")

                    for row in range(row_range):
                        row += 6 + x
                        cell = ws.cell(row=row, column=2)
                        if cell.value is not None:
                            try:
                                if 0 < cell.value < 10 ** 15:
                                    self.calculate(cell.value, row, ws) # row + x
                            except TypeError:
                                pass

                elif checked == self.check_r:
                    row_range = ile * 2 + 3
                    for row in range(row_range):
                        row = row + x
                        if row == 0 + x:
                            ws.cell(row + 6, 2, '\u03A9')
                            ws.merge_cells(f"B{row + 6}:F{row + 6}")
                        elif row == 3 + x:
                            ws.cell(row + 6, 2, 'k\u03A9')
                            ws.merge_cells(f"B{row + 6}:F{row + 6}")
                        elif row == 10 + x:
                            ws.cell(row + 6, 2, 'M\u03A9')
                            ws.merge_cells(f"B{row + 6}:F{row + 6}")
                        elif row < row_range + x - 14:
                            ws.cell(row + 6, 2, zakres1 / 10**7)
                        elif row < row_range + x - 11:
                            ws.cell(row + 6, 2, zakres1 / 10**6)
                        elif row < row_range - 9 + x:
                            ws.cell(row + 6, 2, zakres2/10**8)
                        elif row < row_range - 7 + x:
                            ws.cell(row + 6, 2, zakres2/10**7)
                        elif row < row_range - 4 + x:
                            ws.cell(row + 6, 2, zakres2/10**6)
                        elif row < row_range - 2 + x:
                            ws.cell(row + 6, 2, zakres2/10000000)
                        elif row < row_range + x:
                            ws.cell(row + 6, 2, zakres1/10**6)
                        for col in range(2, 8):
                            cell = ws.cell(row=row+6, column=col)
                            cell.border = border
                            cell.alignment = alignmentCC
                    for row in range(row_range + 3):
                        row = row + 6 + x
                        cell = ws.cell(row=row, column=2)
                        if cell.value is not None:
                            try:
                                if 10 ** 15 > cell.value > 0 == (row - x) % 2 or \
                                        10 ** 15 > cell.value > 0 == (row - (x + 7)) % 2:
                                    ws.merge_cells(f"B{row}:B{row + 1}")
                            except TypeError:
                                pass

                    for row in range(row_range):
                        row += 6 + x
                        cell = ws.cell(row=row, column=2)
                        if cell.value is not None:
                            try:
                                if 0 < cell.value < 10 ** 15:
                                    ws.cell(row + 1, 3, cell.value * 0.9)
                                    ws.cell(row, 3, cell.value * 0.1)
                            except TypeError:
                                pass

        def creat_excelAC(self, ws, checked, ile, zakres1, zakres2):

            col_headers = ['Zakres', 'Punkt Pomiarowy', 'Wartość napięcia odniesienia',
                           'Zmierzona wartość napięcia', 'Poprawka', 'Niepewność pomiaru']

            # ustawianie stylu obramowania dla komórek
            border_style = Side(border_style='thin', color='000000')
            border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

            # ustawianie czcionki i pozycji tekstu
            alignmentCC = Alignment(wrap_text=True, horizontal='center', vertical='center')
            fontA10 = Font(name='Arial', size=10)

            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 13
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 11

            if checked.isChecked():
                x = 10 # 0, tutaj zmienie chyba na wiecej zmienic TEZ W 398, 366

                for i, col in enumerate(col_headers):
                    ws.cell(5 + x, i + 2, col)
                    cell = ws.cell(row=5 + x, column=i+2)
                    cell.border = border
                    cell.font = fontA10
                    cell.alignment = alignmentCC

                if (ile <= 4) and (zakres2 / 100 >= 1) or \
                   (ile == 3) and (zakres2 / 10 >= 1) or \
                   (ile <= 2) and (zakres2 > 1) and (zakres2 < 10):
                    row_range = ile * 3 + 1
                elif (5 <= ile <= 7) and (zakres2 < 1) or \
                     (ile == 6) and (zakres2 / 10 < 1) and (zakres2 > 1):
                    row_range = ile * 3 + 3
                else:
                    row_range = ile * 3 + 2
                if checked != self.check_r:
                    for row in range(row_range):
                        row = row + x
                        try:
                            if (row == 0 + x) and (ile <= 4) and (zakres2 / 100 >= 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 0 + x) and (5 <= ile <= 7) and (zakres2 / 100 >= 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'mV')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'mA')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 4 + x) and (ile == 5) and (zakres2 / 100 >= 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{11 + x}:G{11 + x}")
                            elif (row == 7 + x) and (ile == 6) and (zakres2 / 100 >= 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{15 + x}:G{15 + x}")
                            elif (row == 10 + x) and (ile == 7) and (zakres2 / 100 >= 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{19 + x}:G{19 + x}")
                            elif (row == 0 + x) and (ile <= 3) and (zakres2 / 10 >= 1) and (zakres2 / 100 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 0 + x) and (4 <= ile <= 6) and (zakres2 / 10 >= 1) and (zakres2 / 100 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'mV')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'mA')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 4 + x) and (ile == 4) and (zakres2 / 10 >= 1) and (zakres2 / 100 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{11 + x}:G{11 + x}")
                            elif (row == 7 + x) and (ile == 5) and (zakres2 / 10 >= 1) and (zakres2 / 100 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{15 + x}:G{15 + x}")
                            elif (row == 10 + x) and (ile == 6) and (zakres2 / 10 >= 1) and (zakres2 / 100 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{16 + x}:G{16 + x}")
                            elif (row == 0 + x) and (ile <= 2) and (zakres2 / 10 < 1) and (zakres2 > 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 0 + x) and (3 <= ile <= 5) and (zakres2 / 10 < 1) and (zakres2 > 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'mV')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'mA')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 4 + x) and (ile == 3) and (zakres2 / 10 < 1) and (zakres2 > 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{11 + x}:G{11 + x}")
                            elif (row == 7 + x) and (ile == 4) and (zakres2 / 10 < 1) and (zakres2 > 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{15 + x}:G{15 + x}")
                            elif (row == 10 + x) and (ile == 5) and (zakres2 / 10 < 1) and (zakres2 > 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{19 + x}:G{19 + x}")
                            elif (row == 0 + x) and (ile == 6) and (zakres2 / 10 < 1) and (zakres2 > 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'uV')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'uA')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 4 + x) and (ile == 6) and (zakres2 / 10 < 1) and (zakres2 > 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'mV')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'mA')
                                ws.merge_cells(f"B{11 + x}:G{11 + x}")
                            elif (row == 14 + x) and (ile == 6) and (zakres2 / 10 < 1) and (zakres2 > 1): #bylo 18 chyba 14
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{24 + x}:G{24 + x}")
                            elif (row == 0 + x) and (ile <= 2) and (zakres2 / 10 < 1) and (zakres2 > 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 0 + x) and (ile <= 1) and (zakres2 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 0 + x) and (2 <= ile <= 4) and (zakres2 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'mV')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'mA')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 4 + x) and (ile == 2) and (zakres2 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{11 + x}:G{11 + x}")
                            elif (row == 7 + x) and (ile == 3) and (zakres2 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{15 + x}:G{15 + x}")
                            elif (row == 10 + x) and (ile == 4) and (zakres2 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{19 + x}:G{19 + x}")
                            elif (row == 0 + x) and (5 <= ile <= 6) and (zakres2 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'uV')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'uA')
                                ws.merge_cells(f"B{6 + x}:G{6 + x}")
                            elif (row == 4 + x) and (ile == 5) and (zakres2 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'mV')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'mA')
                                ws.merge_cells(f"B{11 + x}:G{11 + x}")
                            elif (row == 7 + x) and (ile == 6) and (zakres2 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'mV')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'mA')
                                ws.merge_cells(f"B{15 + x}:G{15 + x}")
                            elif (row == 14 + x) and (ile == 5) and (zakres2 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{24 + x}:G{24 + x}")
                            elif (row == 17 + x) and (ile == 6) and (zakres2 < 1):
                                if checked == self.check_dcv or checked == self.check_acv:
                                    ws.cell(row + 6, 2, 'V')
                                elif checked == self.check_dci or checked == self.check_aci:
                                    ws.cell(row + 6, 2, 'A')
                                ws.merge_cells(f"B{28 + x}:G{28 + x}") #git?
                            elif row < row_range - 18 + x:
                                if zakres2 / 10 < 1 < zakres2:
                                    ws.cell(row + 6, 2, zakres2 * 100)
                                elif zakres2 / 10 < 1:
                                    ws.cell(row + 6, 2, zakres2 * 100)
                                else:
                                    ws.cell(row + 6, 2, zakres2 / 10)
                            elif row < row_range - 15 + x:
                                if zakres2 < 1:
                                    ws.cell(row + 6, 2, zakres2 * 1000)
                                else:
                                    ws.cell(row + 6, 2, zakres2)
                            elif row < row_range - 9 + x:
                                if zakres2 / 100 < 1:
                                    ws.cell(row + 6, 2, zakres2 * 10)
                                elif zakres2 < 1:
                                    ws.cell(row + 6, 2, zakres2 * 100)
                                else:
                                    ws.cell(row + 6, 2, zakres2 / 100)
                            elif row < row_range - 6 + x:
                                if zakres2 / 10 < 1:
                                    ws.cell(row + 6, 2, zakres2 * 100)
                                else:
                                    ws.cell(row + 6, 2, zakres2 / 10)
                            elif row < row_range - 3 + x:
                                if zakres2 < 1:
                                    ws.cell(row + 6, 2, zakres2 * 1000)
                                else:
                                    ws.cell(row + 6, 2, zakres2)
                            elif row < row_range + 1 + x:
                                ws.cell(row + 6, 2, zakres1)
                        except AttributeError:
                            pass
                        for col in range(2, 8):
                            cell = ws.cell(row=row+6, column=col)
                            cell.border = border
                            cell.alignment = alignmentCC
                    if ile == 6 and zakres2 / 100 >= 1:
                        end = 1
                    elif ile == 5 and zakres2 / 100 >= 1:
                        end = 2
                    elif ile <= 4 and zakres2 / 100 >= 1:
                        end = 0
                    elif ile == 6 and zakres2 / 10 >= 1:
                        end = 0
                    elif ile == 5 and zakres2 / 10 >= 1:
                        end = 1
                    elif ile == 4 and zakres2 / 10 >= 1:
                        end = 2
                    elif ile == 3 and zakres2 / 10 >= 1:
                        end = 0
                    elif ile <= 4 and zakres2 / 10 >= 1:
                        end = 3
                    elif ile == 3 and zakres2 < 1:
                        end = 1
                    elif ile == 2 and zakres2 < 1:
                        end = 2
                    elif ile == 4 and zakres2 / 10 < 1 < zakres2:
                        end = 1
                    elif ile == 6 and zakres2 < 1:
                        end = 1
                    elif ile == 5 and zakres2 < 1:
                        end = 2
                    elif ile == 6 and zakres2 / 10 < 1 < zakres2:
                        end = 2
                    else:
                        end = 0

                    endy = 0
                    y = 4 * (3 - end)

                    if (ile <= 4) and (zakres2 / 100 >= 1) or \
                            (ile == 3) and (zakres2 / 10 >= 1) or \
                            (ile <= 2) and (zakres2 > 1) and (zakres2 < 10):
                        row_range = ile * 4 + 1
                        for row in range(row_range + 3):
                            row = row + x
                            if (row - (x + 7) >= 0) and (row - (x + 7)) % 3 == 0 and end < 3 or \
                                    (row - (x + 7) >= 0) and (row - (x + 7)) % 3 == 0 and ile == 4 and end <= 3:
                                end += 1
                                ws.merge_cells(f"B{row}:B{row + 2}")
                    elif (5 <= ile <= 7) and (zakres2 < 1) or \
                            (ile == 6) and (zakres2 / 10 < 1) and (zakres2 > 1):
                        row_range = ile * 4 + 3
                        for row in range(row_range + 3):
                            row = row + x
                            if (row - (x + 7) >= 0) and (row - (x + 7)) % 3 == 0 and end < 3 or \
                                    (row - (x + 11) >= 0) and (row - (x + 11)) % 3 == 0 and end < 3:
                                end += 1
                                ws.merge_cells(f"B{row}:B{row + 2}")
                            elif ((100 > row > 60 and row > (6 + x + y)) or 45 > row > (6 + x + y)) and \
                                    (row - x) % 3 == 0 and end >= 3 > endy or \
                                    row > (6 + x + y) and (row > 101 and (row - 102) % 3 == 0) and end >= 3 > endy or \
                                    (60 > row > 45 and (row - (42 + x)) % 3 == 0) and end >= 3 > endy:
                                endy += 1
                                ws.merge_cells(f"B{row}:B{row + 2}")
                            elif (row - (25 + x)) % 3 == 0 and row > y + 19 + x and endy >= 3:
                                ws.merge_cells(f"B{row}:B{row + 2}")
                    else:
                        row_range = ile * 4 + 2
                        for row in range(row_range + 3):
                            row = row + x
                            if (row - (x + 7) >= 0) and (row - (x + 7)) % 3 == 0 and end < 3:
                                end += 1
                                ws.merge_cells(f"B{row}:B{row + 2}")
                            elif ((100 > row > 60 and row > (6 + x + y)) or 45 > row > (6 + x + y)) and \
                                    (row - x) % 3 == 0 and end >= 3 or \
                                    row > (6 + x + y) and (row > 101 and (row-102) % 3 == 0) and end >= 3 or \
                                    (60 > row > 45 and (row - (42 + x)) % 3 == 0) and end >= 3:
                                ws.merge_cells(f"B{row}:B{row + 2}")

                    # for row in range(row_range):
                    #     row += 6 + x
                    #     cell = ws.cell(row=row, column=2)
                    #     if cell.value is not None:
                    #         try:
                    #             if 0 < cell.value < 10 ** 15:
                    #                 self.calculateAC(cell.value, row, ws) # row + x
                    #         except TypeError:
                    #             pass


        def calculate(self, zakres, row, ws):

            ws.cell(row + 3, 3, zakres * (-0.9))
            ws.cell(row + 2, 3, zakres * 0.9)
            ws.cell(row + 1, 3, zakres * 0.5)
            ws.cell(row, 3, zakres * 0.1)

        def calculateAC(self, zakres, row, ws):

            ws.cell(row + 2, 3, zakres * 0.9)
            ws.cell(row + 1, 3, zakres * 0.5)
            ws.cell(row, 3, zakres * 0.1)

        # funkcja do zapisu
        def save_file_dialog(self):

            options = QFileDialog.Options()
            self.saveFileDialog.setDirectory(self.sciezkaSW_zapis.text())
            fileName, _ = self.saveFileDialog.getSaveFileName(self, "Save File",
                                                              self.sciezkaSW_zapis.text(),
                                                              "Excel Files (*.xlsx);;All Files (*)",
                                                              options=options)
            if fileName:
                sender = self.sender()
                # tutaj ustawiam ścieżkę na QLineEdit
                if sender.objectName() == "SzukajSW":
                    path = fileName
                    filename_without_ext, ext = os.path.splitext(path)
                    self.sciezkaSW_zapis.setText(filename_without_ext)
                    print(fileName)

                    # Utwórz nowy plik Excel
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    sheet['A1'] = 'Hello'
                    sheet['B1'] = 'World'
                    # Zapisz plik Excel
                    workbook.save(fileName)

                elif sender.objectName() == "SzukajWynikow":
                    path = fileName
                    filename_without_ext, ext = os.path.splitext(path)
                    self.sciezkaWynik_zapis.setText(filename_without_ext)
                    print(fileName)

                    # Utwórz nowy plik Excel
                    # workbook = openpyxl.Workbook()
                    # sheet = workbook.active
                    # sheet['A1'] = 'Hello'
                    # sheet['B1'] = 'World'

                    wb = Workbook()

                    if self.check_r.isChecked():
                        ws5 = wb.create_sheet("R", 0)
                        self.creat_excel(ws5, self.check_r, self.ilR.value(),
                                         self.zakresR.value(), self.zakresR.value())
                    if self.check_aci.isChecked():
                        ws4 = wb.create_sheet("ACI", 0)
                        self.creat_excel(ws4, self.check_aci, self.ilACI.value(),
                                         self.zakresACI.value(), self.zakresACI_2.value())
                    if self.check_dci.isChecked():
                        ws3 = wb.create_sheet("DCI", 0)
                        self.creat_excel(ws3, self.check_dci, self.ilDCI.value(),
                                         self.zakresDCI.value(), self.zakresDCI_2.value())
                    if self.check_acv.isChecked():
                        ws2 = wb.create_sheet("ACV", 0)
                        self.creat_excel(ws2, self.check_acv, self.ilACV.value(),
                                         self.zakresACV.value(), self.zakresACV_2.value())
                    if self.check_dcv.isChecked():
                        ws1 = wb.create_sheet("DCV", 0)
                        self.creat_excel(ws1, self.check_dcv, self.ilDCV.value(),
                                         self.zakresDCV.value(), self.zakresDCV_2.value())

                    self.zapis_pliku(self.sciezkaWynik_zapis.text(), wb)

                    # Zapisz plik Excel
                    # workbook.save(fileName)


        def zapis_pliku(self, path, wb):

            # Zapisanie pliku
            if os.path.exists(path + ".xlsx"):
                # Wyświetlenie okna dialogowego z pytaniem o nadpisanie pliku
                msg_box = QMessageBox()
                msg_box.setText("Plik już istnieje. Czy chcesz go nadpisać?")
                msg_box.setWindowTitle("Nadpisać plik?")
                msg_box.setIcon(QMessageBox.Question)
                msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
                msg_box.setDefaultButton(QMessageBox.No)
                # Obsługa wyboru użytkownika
                choice = msg_box.exec()
                if choice == QMessageBox.Yes:
                    try:
                        os.remove(path + ".xlsx")
                        print("Plik już istnieje, nadpisywanie...")
                        wb.save(path + ".xlsx")
                        self.open_folder(path)
                    except PermissionError as e:
                        msg = QMessageBox()
                        msg.setIcon(QMessageBox.Critical)
                        msg.setText("Błąd zapisu pliku")
                        msg.setInformativeText(str(e))
                        msg.setWindowTitle("Błąd")
                        msg.exec()
            else:
                print("Plik nie istnieje, zapisywanie...")
                wb.save(path + ".xlsx")
                self.open_folder(path)

        def naglowek(self, ws):

            ws.merge_cells("A2:B2")
            ws.merge_cells("C2:G2")
            ws.merge_cells("A3:G3")
            ws.merge_cells("A4:B4")
            #ws.merge_cells("B4:C4")
            ws.merge_cells("D4:F4")

            # ustawianie stylu obramowania dla komórek
            border_style = Side(border_style='thin', color='000000')
            border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

            # ustawienie obramowania dla komórki
            for row in range(2, 5):
                for column in range(1, 8):
                    cell = ws.cell(row=row, column=column)
                    cell.border = border

            ws.row_dimensions[2].height = 92
            ws.row_dimensions[3].height = 41.25
            ws.row_dimensions[4].height = 30
            ws.column_dimensions['A'].width = 13
            ws.column_dimensions['G'].width = 17
            ws.column_dimensions['H'].width = 0


            # ustawianie czcionki i pozycji tekstu
            alignmentCT = Alignment(wrap_text=True, horizontal='center', vertical='top')
            alignmentCC = Alignment(wrap_text=True, horizontal='center', vertical='center')
            alignmentLC = Alignment(wrap_text=True, horizontal='left', vertical='center')
            alignmentLT = Alignment(wrap_text=True, horizontal='left', vertical='top')
            fontA12 = Font(name='Arial', size=12)
            fontA22 = Font(name='Arial', size=22)
            fontA10 = Font(name='Arial', size=10)
            fontA10B = Font(name='Arial', size=10, bold=True)

            # wstawienie loga
            img = Image('image/logov3.png')
            img.height = 116
            img.width = 140
            ws.add_image(img, 'A2')

            ws['C2'] = ('Metro-Lab Agnieszka Greń \n '
                        'Warszawa 01-386 ul. Reżyserska 5 \n '
                        'tel. (+48) 661 472 870 \n '
                        'e-mail: poczta@metro-lab.pl \n'
                        'www.metro-lab.pl')

            ws['C2'].alignment = alignmentCC
            ws['C2'].font = fontA12

            ws['A3'] = "ŚWIADECTWO WZORCOWANIA"
            ws['A3'].font = fontA22
            ws['A3'].alignment = alignmentCC

            ws['A4'] = "Data wydania:"
            ws['A4'].font = fontA10
            ws['A4'].alignment = alignmentLC

            ws['C4'] = self.data_wzorcowania.text()
            ws['C4'].font = fontA10
            ws['C4'].alignment = alignmentLC

            ws['D4'] = "Nr świadectwa: " + self.numer_swiadectwa.text()
            ws['D4'].font = fontA10
            ws['D4'].alignment = alignmentLC

            ws['G4'] = "Strona 1/6"  # ogarnac str
            ws['G4'].font = fontA10
            ws['G4'].alignment = alignmentLC

            if ws.title == "DCV":
                ws['A7'] = "        1. Oględziny zewnętrzne:"
                ws['A8'] = "           a. stan techniczny obudowy, gniazd i przełączników - Uwag brak."
                ws['A9'] = "        2. Sprawdzenie wstępne "
                ws['A10'] = "           a. działanie elementów służących do regulacji i przełączania - Uwag brak"
                ws['A12'] = "        Wyznaczenie poprawek wskazań napięcia stałego DC."
            if ws.title == "ACV":
                ws['A12'] = "        Wyznaczenie poprawek wskazań napięcia przemiennego AC."
                ws['A13'] = "        f = 60 Hz"
            if ws.title == "ACI":
                ws['A12'] = "        Wyznaczenie poprawek wskazań nateżenia prądu przemiennego AC."
                ws['A13'] = "        f = 60 Hz"
            if ws.title == "DCI":
                ws['A12'] = "        Wyznaczenie poprawek wskazań nateżenia prądu stałego DC."
            if ws.title == "R":
                ws['A12'] = "        Wyznaczenie poprawek wskazań rezystancji."


        def swiadectwo_final(self):

            try:

                self.swiadectwo()

                # START czesci do kiopiowania i tworzenia 1 pliku z danymi calymi

                # pierwszy plik Excel
                path1 = self.sciezkaSW_zapis.text() + ".xlsx"
                # drugi plik Excel
                path2 = self.sciezkaWynik_zapis.text() + ".xlsx"

                wb1 = xw.Book(path2)
                wb2 = xw.Book(path1)

                for i in range(wb1.sheets.count - 1):
                    ws1 = wb1.sheets(i+1)
                    ws1.api.Copy(After=wb2.sheets(i+1).api)

                wb2.save()
                wb2.app.quit()

                # KONIEC

                # START robienie operacji na głównym swiadectwie

                path = self.sciezkaSW_zapis.text() + ".xlsx"
                wb = openpyxl.load_workbook(path)

                worksheets = wb.sheetnames

                for i, sheet in enumerate(worksheets):
                    if sheet != "Swiadectwo":
                        if sheet != "Sheet":
                            ws = wb[sheet]
                            ws.delete_cols(3)
                            self.naglowek(ws)
                            ws['G4'] = f"Strona {i+1}/{len(worksheets)-1}"
                            ws.oddFooter.center.text = "Niniejsze świadectwo może być okazywane lub kopiowane tylko w całości."
                    elif sheet == "Swiadectwo":
                        ws = wb[sheet]
                        ws['E4'] = f"Strona {i + 1}/{len(worksheets) - 1}"

                # zapis pliku
                wb.save(path)
            except Exception as e:
                self.error2.setText(str(e))
                print(e)
                pass

        # generowanie świadectwa
        def swiadectwo(self):

            # Tworzenie nowego pliku Excel
            wb = Workbook()
            # Tworzenie nowego arkusza
            # ws = wb.active
            ws = wb.create_sheet("Swiadectwo", 0)

            # Połączenie komórek
            for row in range(7, 15):
                ws.merge_cells(f"C{row}:E{row}")

            for col in ['B', 'C', 'D', 'E']:
                ws.merge_cells(f'{col}4:{col}5')

            ws.merge_cells("C2:E2")
            ws.merge_cells("B3:E3")
            ws.merge_cells("B15:E19")
            # ws.merge_cells("A4:D5")

            # ustawianie stylu obramowania dla komórek
            border_style = Side(border_style='thin', color='000000')
            border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

            # ustawienie obramowania dla komórki
            for row in range(2, 20):
                if row != 6:
                    for column in range(2, 6):
                        cell = ws.cell(row=row, column=column)
                        cell.border = border

            # ustawienie wielkośći komórek
            ws.row_dimensions[2].height = 90
            ws.row_dimensions[3].height = 41.25
            ws.row_dimensions[4].height = 15
            ws.row_dimensions[5].height = 15
            ws.row_dimensions[6].height = 15
            ws.row_dimensions[7].height = 53.25
            ws.row_dimensions[8].height = 45.75
            ws.row_dimensions[9].height = 37.50
            ws.row_dimensions[10].height = 48
            ws.row_dimensions[11].height = 37.50
            ws.row_dimensions[12].height = 79.50
            ws.row_dimensions[13].height = 108.75
            ws.row_dimensions[14].height = 37.50
            ws.row_dimensions[19].height = 21

            ws.column_dimensions['A'].width = 1
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 13.71
            ws.column_dimensions['D'].width = 30.57
            ws.column_dimensions['E'].width = 20.85

            # ustawianie czcionki i pozycji tekstu
            alignmentCT = Alignment(wrap_text=True, horizontal='center', vertical='top')
            alignmentCC = Alignment(wrap_text=True, horizontal='center', vertical='center')
            alignmentLC = Alignment(wrap_text=True, horizontal='left', vertical='center')
            alignmentLT = Alignment(wrap_text=True, horizontal='left', vertical='top')
            fontA12 = Font(name='Arial', size=12)
            fontA22 = Font(name='Arial', size=22)
            fontA10 = Font(name='Arial', size=10)
            fontA10B = Font(name='Arial', size=10, bold=True)

            # wstawienie loga
            img = Image('image/logov3.png')
            img.height = 116
            img.width = 140
            ws.add_image(img, 'B2')

            ws['C2'] = ('Metro-Lab Agnieszka Greń \n '
                        'Warszawa 01-386 ul. Reżyserska 5 \n '
                        'tel. (+48) 661 472 870 \n '
                        'e-mail: poczta@metro-lab.pl \n'
                        'www.metro-lab.pl')

            ws['C2'].alignment = alignmentCC
            ws['C2'].font = fontA12

            ws['B3'] = "ŚWIADECTWO WZORCOWANIA"
            ws['B3'].font = fontA22
            ws['B3'].alignment = alignmentCC

            ws['B4'] = "Data wydania:"
            ws['B4'].font = fontA10
            ws['B4'].alignment = alignmentLC

            ws['C4'] = self.data_wzorcowania.text()
            ws['C4'].font = fontA10
            ws['C4'].alignment = alignmentLC

            ws['D4'] = "Nr świadectwa: " + self.numer_swiadectwa.text()
            ws['D4'].font = fontA10
            ws['D4'].alignment = alignmentLC

            ws['E4'] = "Strona 1/6" #ogarnac str
            ws['E4'].font = fontA10
            ws['E4'].alignment = alignmentLC

            ws['B7'] = "PRZEDMIOT \nWZORCOWANIA"
            ws['B8'] = "ZGŁASZAJĄCY"
            ws['B9'] = "METODA \nWZORCOWANIA"
            ws['B10'] = "WARUNKI \nŚRODOWISKOWE"
            ws['B11'] = "DATA WYKONANIA \nWZORCOWANIA"
            ws['B12'] = "SPÓJNOŚĆ \nPOMIAROWA"
            ws['B13'] = "NIEPEWNOŚĆ \nPOMIARU"
            ws['B14'] = "WYNIKI \nWZOROCWANIA"

            ws['C7'] = self.przedmiot_wzorcowania.toPlainText()
            ws['C8'] = self.Zglaszajacy.toPlainText()
            ws['C9'] = self.metoda_wzorcowania.toPlainText()
            ws['C10'] = self.warunki_srodowiskowe.toPlainText()
            ws['C11'] = self.data_wzorcowania.text()
            ws['C12'] = self.spojnosc_pomiarowa.toPlainText()
            ws['C13'] = self.niepewnosc_pomiaru.toPlainText()
            ws['C14'] = "Podano na kolejnych stronach niniejszego świadectwa"
            ws['B15'] = "                                               Robert Greń"


            # iteracja po wierszach i kolumnach, zmiana stylu komórek
            for row in range(6, 20):
                    cell = ws.cell(row=row, column=3)
                    cell.font = fontA10
                    cell.alignment = alignmentLC

            for row in range(6, 20):
                    cell = ws.cell(row=row, column=2)
                    cell.font = fontA10B
                    cell.alignment = alignmentLC

            ws['B15'].alignment = alignmentCC
            ws['B15'].font = fontA10

            ws.oddFooter.center.text = "Niniejsze świadectwo może być okazywane lub kopiowane tylko w całości."

            # zapis pliku
            self.zapis_pliku(self.sciezkaSW_zapis.text(), wb)




    app = QApplication(sys.argv)

    window = MainWindow()
    window.show()
    app.exec()



